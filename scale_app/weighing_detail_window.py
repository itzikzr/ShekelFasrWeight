"""
חלון פירוט שקילה — כל הקריאות הגולמיות של שקילה בודדת, עם סימון ירוק לקריאות
שמהן חושב המשקל בפועל (החלון השטוח שבחר decide_weight), וסיכום החישוב מתחת לטבלה.

נשמר בפירוט רק ל-db.DETAIL_RETENTION_COUNT השקילות האחרונות (ראו db.py) — לשקילות
ישנות יותר הטבלה תוצג ריקה, עם הודעה מתאימה במקום.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from . import db, theme
from .formatting import fmt_weight

VERDICT_LABELS = {"green": "ירוק — בטווח", "red": "אדום — מעל", "yellow": "צהוב — מתחת",
                  "none": "ללא מוצר"}


class WeighingDetailWindow(tk.Toplevel):
    def __init__(self, main, weighing_id: int):
        super().__init__(main.root)
        self.weighing_id = weighing_id
        self.title(f"פירוט שקילה #{weighing_id}")
        self.geometry("540x680")
        self.configure(background=theme.BG)

        self.weighing = weighing = db.get_weighing(weighing_id)
        self.readings = readings = db.get_weighing_readings(weighing_id)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=10, pady=10)
        body.rowconfigure(0, weight=1)
        body.columnconfigure(0, weight=1)

        cols = ("#", "זמן (s)", "משקל (kg)")
        self.tree = ttk.Treeview(body, columns=cols, show="headings")
        for c, w in zip(cols, (50, 110, 120)):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center")
        self.tree.grid(row=0, column=0, sticky="nsew")

        sb = ttk.Scrollbar(body, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        sb.grid(row=0, column=1, sticky="ns")

        self.tree.tag_configure("inwin", background=theme.GREEN_BG, foreground=theme.GREEN)

        if not readings:
            ttk.Label(self, text="לא נשמרו קריאות מפורטות לשקילה זו — נשמר פירוט רק "
                                 f"ל-{db.DETAIL_RETENTION_COUNT} השקילות האחרונות.",
                      foreground=theme.TEXT_MUTED, wraplength=500,
                      justify="right").pack(padx=10, pady=(0, 6))
        else:
            for i, r in enumerate(readings, start=1):
                tag = ("inwin",) if r["in_window"] else ()
                self.tree.insert("", "end", tags=tag,
                                 values=(i, f"{r['t_offset']:.3f}",
                                         fmt_weight(r["weight"], force_sign=True)))
            ttk.Label(self, text="ירוק = הקריאות שמהן חושב המשקל הסופי (החלון השטוח ביותר)",
                      foreground=theme.GREEN, wraplength=500,
                      justify="right").pack(padx=10, pady=(0, 6))

        # ── סיכום החישוב ──
        summary = ttk.LabelFrame(self, text="נתוני החישוב")
        summary.pack(fill="x", padx=10, pady=(0, 10))
        pad = {"padx": 8, "pady": 3}

        if weighing is None:
            ttk.Label(summary, text="השקילה לא נמצאה — ייתכן שנמחקה.").grid(
                row=0, column=0, columnspan=2, **pad)
        else:
            target_str = (fmt_weight(weighing["target_weight"])
                          if weighing["target_weight"] is not None else "—")
            elapsed_str = (f"{weighing['elapsed_seconds']:.2f}s"
                          if weighing["elapsed_seconds"] is not None else "—")
            rows = [
                ("משקל שנקבע:", f"{fmt_weight(weighing['decided_weight'], force_sign=True)} kg"),
                ("בסיס החישוב:", weighing["basis"] or "—"),
                ("מוצר:", weighing["product_name"] or "— ללא מוצר —"),
                ("משקל מטרה:", target_str),
                ("תוצאה:", VERDICT_LABELS.get(weighing["verdict"], weighing["verdict"])),
                ("מין / מקס:", f"{fmt_weight(weighing['min_weight'], force_sign=True)} / "
                              f"{fmt_weight(weighing['max_weight'], force_sign=True)}"),
                ("קריאות:", str(weighing["reading_count"] or 0)),
                ("משך:", elapsed_str),
                ("זמן:", weighing["timestamp"]),
            ]
            for i, (label, value) in enumerate(rows):
                ttk.Label(summary, text=label,
                         style="Muted.TLabel").grid(row=i, column=1, sticky="e", **pad)
                ttk.Label(summary, text=value,
                         font=("TkDefaultFont", 10, "bold")).grid(row=i, column=0, sticky="e",
                                                                   **pad)

        btns = ttk.Frame(self)
        btns.pack(pady=(0, 10))
        ttk.Button(btns, text="ייצוא לאקסל", style="Accent.TButton",
                  command=self._export_excel).pack(side="left", padx=6)
        ttk.Button(btns, text="סגור", command=self.destroy).pack(side="left", padx=6)
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror("שגיאה", "ייצוא לאקסל דורש את הספרייה openpyxl.\n"
                                          "הרץ: pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")],
            initialfile=f"שקילה_{self.weighing_id}.xlsx")
        if not path:
            return

        wb = openpyxl.Workbook()
        header_font = Font(bold=True)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        right = Alignment(horizontal="right")

        ws = wb.active
        ws.title = "קריאות"
        ws.sheet_view.rightToLeft = True
        ws.append(["#", "זמן (s)", "משקל (kg)", "בחלון החישוב"])
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = right
        for i, r in enumerate(self.readings, start=1):
            ws.append([i, round(r["t_offset"], 3), r["weight"], "כן" if r["in_window"] else ""])
            if r["in_window"]:
                for cell in ws[ws.max_row]:
                    cell.fill = green_fill
        for col, width in zip("ABCD", (6, 12, 14, 14)):
            ws.column_dimensions[col].width = width

        ws2 = wb.create_sheet("סיכום")
        ws2.sheet_view.rightToLeft = True
        w = self.weighing
        if w is not None:
            target_str = fmt_weight(w["target_weight"]) if w["target_weight"] is not None else "—"
            elapsed_str = f"{w['elapsed_seconds']:.2f}s" if w["elapsed_seconds"] is not None else "—"
            summary_rows = [
                ("שקילה #", self.weighing_id),
                ("זמן", w["timestamp"]),
                ("מוצר", w["product_name"] or "— ללא מוצר —"),
                ("משקל שנקבע (kg)", w["decided_weight"]),
                ("בסיס החישוב", w["basis"] or "—"),
                ("משקל מטרה (kg)", target_str),
                ("תוצאה", VERDICT_LABELS.get(w["verdict"], w["verdict"])),
                ("מינימום (kg)", w["min_weight"]),
                ("מקסימום (kg)", w["max_weight"]),
                ("קריאות", w["reading_count"] or 0),
                ("משך", elapsed_str),
            ]
            for label, value in summary_rows:
                ws2.append([label, value])
            for row in ws2.iter_rows():
                row[0].font = header_font
                for cell in row:
                    cell.alignment = right
            ws2.column_dimensions["A"].width = 20
            ws2.column_dimensions["B"].width = 22

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))
            return
        messagebox.showinfo("ייצוא הושלם", f"נשמר אל:\n{path}")
