"""
דוחות — סטטיסטיקת תקינות וגרף מגמת משקל לפי מוצר וטווח זמן, עם ייצוא לאקסל.

הגרף (widgets.TrendChart) מוצג תמיד, גם כש"כל המוצרים" נבחר — רק שאז אין קו
מטרה/רצועת טולרנס (אלה שייכים למוצר ספציפי). מוגבל ל-CHART_MAX_POINTS כדי
שטווח "כל הזמן" על היסטוריה גדולה לא יצייר אלפי נקודות בלתי קריאות — כשיש
חיתוך מוצגת הערה על כך (בלי חיתוך שקט), ראו refresh().
"""

import tkinter as tk
from datetime import datetime, timedelta
from tkinter import ttk, filedialog, messagebox

from . import db, rtl, theme
from .formatting import fmt_weight
from .widgets import DatePicker, TrendChart

ALL_LABEL = "כל המוצרים"
VERDICT_LABELS = {"green": "ירוק — בטווח", "red": "אדום — מעל", "yellow": "צהוב — מתחת",
                  "none": "ללא מוצר"}

RANGE_FILTERS = [
    ("today", "היום"),
    ("week", "7 ימים אחרונים"),
    ("month", "30 ימים אחרונים"),
    ("all", "כל הזמן"),
    ("custom", "טווח מותאם (מ-עד)"),
]

CHART_MAX_POINTS = 300


class ReportsWindow(tk.Toplevel):
    def __init__(self, main):
        super().__init__(main.root)
        self.main = main
        self.title("דוחות")
        self.minsize(900, 640)
        theme.maximize_window(self)
        self.configure(background=theme.BG)

        self._current_rows = []
        # Combobox מציג טקסט מהופך-bidi בלינוקס (ראו rtl.py) — הבחירה הנקראת
        # בחזרה מתורגמת למפתח/שם המקורי לפני שימוש, בדיוק כמו history_window.
        self._product_display_to_name = {}
        self._range_display_to_key = {rtl.visual(label): key for key, label in RANGE_FILTERS}

        filters = ttk.Frame(self)
        filters.pack(fill="x", padx=12, pady=(12, 6))

        ttk.Label(filters, text="מוצר:").pack(side="right", padx=(0, 4))
        self.product_var = tk.StringVar()
        self.product_combo = ttk.Combobox(filters, textvariable=self.product_var,
                                          state="readonly", width=22)
        self.product_combo.pack(side="right", padx=(0, 16))
        self.product_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        ttk.Label(filters, text="טווח:").pack(side="right", padx=(0, 4))
        range_values = [label for _, label in RANGE_FILTERS]
        self.range_var = tk.StringVar(value=rtl.visual(RANGE_FILTERS[1][1]))  # "7 ימים אחרונים"
        self.range_combo = ttk.Combobox(filters, textvariable=self.range_var,
                                        state="readonly", width=18,
                                        values=[rtl.visual(v) for v in range_values])
        self.range_combo.pack(side="right", padx=(0, 16))
        self.range_combo.bind("<<ComboboxSelected>>", lambda e: self._on_range_changed())

        # שדות "מ-עד" — נראים רק כש"טווח מותאם" נבחר, ראו _on_range_changed.
        self.custom_range_frame = ttk.Frame(filters)
        ttk.Label(self.custom_range_frame, text="מתאריך:").pack(side="right", padx=(0, 4))
        self.date_from_picker = DatePicker(self.custom_range_frame, on_change=self.refresh)
        self.date_from_picker.pack(side="right", padx=(0, 16))
        ttk.Label(self.custom_range_frame, text="עד תאריך:").pack(side="right", padx=(0, 4))
        self.date_to_picker = DatePicker(self.custom_range_frame, on_change=self.refresh)
        self.date_to_picker.pack(side="right")

        ttk.Button(filters, text="ייצוא לאקסל", style="Accent.TButton",
                  command=self._export_excel).pack(side="left")

        range_display_row = ttk.Frame(self)
        range_display_row.pack(fill="x", padx=12, pady=(0, 6))
        self.range_display_var = tk.StringVar()
        ttk.Label(range_display_row, textvariable=self.range_display_var,
                 style="Muted.TLabel").pack(side="right")

        self.summary_frame = ttk.Frame(self)
        self.summary_frame.pack(fill="x", padx=12, pady=(0, 10))
        self._build_summary_cards()

        chart_card = ttk.Frame(self, style="Card.TFrame")
        chart_card.pack(fill="both", expand=True, padx=12, pady=(0, 12))
        self.chart = TrendChart(chart_card)
        self.chart.pack(fill="both", expand=True, padx=8, pady=(8, 2))
        self.note_var = tk.StringVar()
        ttk.Label(chart_card, textvariable=self.note_var, style="Card.Muted.TLabel",
                 wraplength=760, justify="right").pack(padx=8, pady=(0, 8), anchor="e")

        self.refresh()
        self.protocol("WM_DELETE_WINDOW", self.destroy)

    def refresh_theme(self):
        self.configure(background=theme.BG)
        self.chart.refresh_theme()

    def _build_summary_cards(self):
        tiles = [("total", "סה\"כ שקילות"), ("green", "ירוק"), ("red", "אדום"),
                 ("yellow", "צהוב"), ("avg", "משקל ממוצע"), ("range", "טווח מין–מקס")]
        self._summary_vars = {}
        for key, title in tiles:
            card = ttk.Frame(self.summary_frame, style="Card.TFrame")
            card.pack(side="right", padx=4, fill="both", expand=True)
            ttk.Label(card, text=title, style="Card.Muted.TLabel").pack(padx=10, pady=(8, 2))
            var = tk.StringVar(value="—")
            self._summary_vars[key] = var
            ttk.Label(card, textvariable=var, style="Card.TLabel",
                     font=("TkDefaultFont", 14, "bold")).pack(padx=10, pady=(0, 8))

    def _refresh_product_list(self):
        names = [ALL_LABEL] + [p["name"] for p in db.list_products()]
        current_display = self.product_var.get()
        current_name = self._product_display_to_name.get(current_display, current_display)

        display_names = [rtl.visual(n) for n in names]
        self._product_display_to_name = dict(zip(display_names, names))
        self.product_combo.config(values=display_names)
        if current_name not in names:
            # ברירת מחדל: המוצר הראשון אם קיים, כדי שהגרף יציג קו מטרה/טולרנס
            # מרגע הפתיחה הראשונה — לא "כל המוצרים" סתמי בלי שום קו ייחוס.
            default_name = names[1] if len(names) > 1 else ALL_LABEL
            self.product_var.set(rtl.visual(default_name))

    def _on_range_changed(self):
        range_display = self.range_var.get()
        range_key = self._range_display_to_key.get(range_display, "week")
        if range_key == "custom":
            self.custom_range_frame.pack(side="right", padx=(0, 16))
        else:
            self.custom_range_frame.pack_forget()
        self.refresh()

    def _since_until_bound(self, key):
        """
        מחזיר (since, until) בפורמט timestamp מחרוזתי של ה-DB. עבור הטווחים
        הקבועים until הוא None (עד "עכשיו"); "טווח מותאם" קורא את
        DatePicker.get() (תמיד "" או "YYYY-MM-DD" תקין — נבחר מלוח שנה, לא
        מוקלד, אז אין צורך באימות פורמט) — until כולל את כל יום ה"עד"
        (23:59:59), לא רק חצות שלו.
        """
        now = datetime.now()
        if key == "today":
            return now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat(
                timespec="seconds"), None
        if key == "week":
            return (now - timedelta(days=7)).isoformat(timespec="seconds"), None
        if key == "month":
            return (now - timedelta(days=30)).isoformat(timespec="seconds"), None
        if key == "custom":
            from_text = self.date_from_picker.get()
            to_text = self.date_to_picker.get()
            since = f"{from_text}T00:00:00" if from_text else None
            until = f"{to_text}T23:59:59" if to_text else None
            return since, until
        return None, None   # "all"

    def _resolve_display_range(self, key, since, rows_desc):
        """ מחרוזת "מציג נתונים: מ – עד" לתצוגה — לכל הטווחים, לא רק "מותאם",
        כדי שהמשתמש יראה את התאריכים הקונקרטיים שמאחורי "7 ימים אחרונים" וכו'. """
        today = datetime.now().date().isoformat()
        if key == "custom":
            frm = self.date_from_picker.get() or "תחילת ההיסטוריה"
            to = self.date_to_picker.get() or today
            return frm, to
        if key == "all":
            frm = rows_desc[-1]["timestamp"][:10] if rows_desc else today
            return frm, today
        frm = since[:10] if since else today
        return frm, today

    @staticmethod
    def _chart_label(timestamp):
        # "YYYY-MM-DDTHH:MM:SS" -> "MM-DD HH:MM", קצר מספיק לתוויות ציר X.
        if not timestamp or len(timestamp) < 16:
            return timestamp or ""
        return f"{timestamp[5:10]} {timestamp[11:16]}"

    def _update_summary(self, rows):
        counts = {"green": 0, "red": 0, "yellow": 0, "none": 0}
        weights = []
        for r in rows:
            counts[r["verdict"]] = counts.get(r["verdict"], 0) + 1
            weights.append(r["decided_weight"])
        total = len(rows)

        def pct(n):
            return f" ({n / total * 100:.0f}%)" if total else ""

        self._summary_vars["total"].set(str(total))
        self._summary_vars["green"].set(f"{counts['green']}{pct(counts['green'])}")
        self._summary_vars["red"].set(f"{counts['red']}{pct(counts['red'])}")
        self._summary_vars["yellow"].set(f"{counts['yellow']}{pct(counts['yellow'])}")
        if weights:
            self._summary_vars["avg"].set(f"{fmt_weight(sum(weights) / len(weights))} kg")
            self._summary_vars["range"].set(
                f"{fmt_weight(min(weights))}–{fmt_weight(max(weights))} kg")
        else:
            self._summary_vars["avg"].set("—")
            self._summary_vars["range"].set("—")

    def refresh(self):
        self._refresh_product_list()

        product = None
        display = self.product_var.get()
        name = self._product_display_to_name.get(display, display)
        if name != ALL_LABEL:
            for p in db.list_products():
                if p["name"] == name:
                    product = p
                    break

        range_display = self.range_var.get()
        range_key = self._range_display_to_key.get(range_display, "week")
        since, until = self._since_until_bound(range_key)

        rows_desc = db.list_weighings(product_id=(product["id"] if product else None),
                                      since=since, until=until)
        self._current_rows = rows_desc
        self._update_summary(rows_desc)

        frm, to = self._resolve_display_range(range_key, since, rows_desc)
        self.range_display_var.set(f"מציג נתונים: {frm} – {to}")

        rows_asc = list(reversed(rows_desc))
        truncated = len(rows_asc) > CHART_MAX_POINTS
        if truncated:
            rows_asc = rows_asc[-CHART_MAX_POINTS:]

        points = [(self._chart_label(r["timestamp"]), r["decided_weight"], r["verdict"])
                 for r in rows_asc]

        if product is not None:
            self.chart.set_data(points, target=product["target_weight"],
                               tolerance_upper=product["tolerance_upper"],
                               tolerance_lower=product["tolerance_lower"],
                               empty_message="אין שקילות למוצר זה בטווח הנבחר")
        else:
            self.chart.set_data(points, empty_message="אין שקילות בטווח הנבחר")

        if truncated:
            self.note_var.set(f"מוצגות ב-{CHART_MAX_POINTS} הנקודות האחרונות מתוך "
                              f"{len(rows_desc)} שקילות בטווח הנבחר — הסטטיסטיקה מעל "
                              f"מחושבת על כל {len(rows_desc)} השקילות.")
        elif product is None and rows_asc:
            self.note_var.set("נבחר \"כל המוצרים\" — הגרף מציג את המגמה בלי קו מטרה/רצועת "
                              "טולרנס (אלה שייכים למוצר ספציפי); לצפייה בהם בחר מוצר.")
        else:
            self.note_var.set("")

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            messagebox.showerror("שגיאה", "ייצוא לאקסל דורש את הספרייה openpyxl.\n"
                                          "הרץ: pip install openpyxl")
            return

        path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel", "*.xlsx")],
                                            initialfile="דוח_שקילות.xlsx")
        if not path:
            return

        wb = openpyxl.Workbook()
        header_font = Font(bold=True)
        right = Alignment(horizontal="right")

        ws = wb.active
        ws.title = "סיכום"
        ws.sheet_view.rightToLeft = True
        product_display = self.product_var.get()
        product_name = self._product_display_to_name.get(product_display, product_display)
        summary_rows = [
            ("מוצר", product_name),
            ("טווח", self.range_var.get()),
            ("תאריכים", self.range_display_var.get()),
            ("סה\"כ שקילות", self._summary_vars["total"].get()),
            ("ירוק", self._summary_vars["green"].get()),
            ("אדום", self._summary_vars["red"].get()),
            ("צהוב", self._summary_vars["yellow"].get()),
            ("משקל ממוצע", self._summary_vars["avg"].get()),
            ("טווח מין–מקס", self._summary_vars["range"].get()),
        ]
        for label, value in summary_rows:
            ws.append([label, value])
        for row in ws.iter_rows():
            row[0].font = header_font
            for cell in row:
                cell.alignment = right
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 26

        ws2 = wb.create_sheet("שקילות")
        ws2.sheet_view.rightToLeft = True
        ws2.append(["#", "זמן", "מוצר", "משקל (kg)", "מטרה", "תוצאה", "קריאות", "משך"])
        for cell in ws2[1]:
            cell.font = header_font
            cell.alignment = right
        for r in self._current_rows:
            target_str = fmt_weight(r["target_weight"]) if r["target_weight"] is not None else "—"
            elapsed_str = f"{r['elapsed_seconds']:.2f}s" if r["elapsed_seconds"] is not None else "—"
            ws2.append([r["id"], r["timestamp"], r["product_name"] or "— ללא מוצר —",
                       r["decided_weight"], target_str,
                       VERDICT_LABELS.get(r["verdict"], r["verdict"]),
                       r["reading_count"] or 0, elapsed_str])
        for col, width in zip("ABCDEFGH", (6, 20, 22, 12, 10, 16, 10, 10)):
            ws2.column_dimensions[col].width = width

        try:
            wb.save(path)
        except Exception as e:
            messagebox.showerror("שגיאה", str(e))
            return
        messagebox.showinfo("ייצוא הושלם", f"נשמר אל:\n{path}")
